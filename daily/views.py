from django.shortcuts import render, redirect
from openpyxl import load_workbook, Workbook
from .forms import TaskForm
import os

EXCEL_FILE = 'tasks.xlsx'

# Create the Excel file with headers if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Flagship Project", "Task", "Duration", "Date", "Notes", "Comment"])
    wb.save(EXCEL_FILE)

def read_tasks():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    return list(ws.iter_rows(min_row=2, values_only=True))

def task_form(request, row_id=None):
    tasks = read_tasks()
    if request.method == 'POST':
        form = TaskForm(request.POST)
        if form.is_valid():
            data = [
                form.cleaned_data['flagship'],
                form.cleaned_data['task'],
                form.cleaned_data['duration'],
                form.cleaned_data['date'].strftime('%Y-%m-%d'),
                form.cleaned_data['notes'],
                form.cleaned_data['comment']
            ]

            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            if row_id is not None:
                for i, _ in enumerate(data):
                    ws.cell(row=row_id+2, column=i+1, value=data[i])
            else:
                ws.append(data)
            wb.save(EXCEL_FILE)
            return redirect('task_form')
    else:
        if row_id is not None:
            selected = tasks[row_id]
            form = TaskForm(initial={
                'flagship': selected[0],
                'task': selected[1],
                'duration': selected[2],
                'date': selected[3],
                'notes': selected[4],
                'comment': selected[5],
            })
        else:
            form = TaskForm()

    return render(request, 'daily/task_form.html', {'form': form, 'task_data': enumerate(tasks)})

def delete_task(request, row_id):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.delete_rows(row_id + 2)  # +2 to skip header and 0-based index
    wb.save(EXCEL_FILE)
    return redirect('task_form')


